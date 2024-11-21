import pandas as pd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

file_path = r'D:\WorkSpace\Fatigue-Driven-Detection-Based-on-CNN-master\data\perclos.xlsx'
col = pd.read_excel(file_path, usecols=[1])
list = col.values.tolist()
perlist = []
for s in list:
    perlist.append(s[0])
# print(len(perlist))

# 打开现有的Excel文件
workbook = openpyxl.load_workbook(file_path)
# 获取第一个工作表
sheet = workbook.active  # 或者使用 sheet = workbook['Sheet1']
# 向指定单元格写入数据
# sheet.cell(row=0, column=2).value = 'perclos'

'''
for index in range(1,30):
    sheet.cell(row=index, column=3).value = 0

for index in range(30,len(perlist)+1):
    # 使用切片获取需要的元素
    calc = perlist[index-29:index+1]
    perclos=1-np.average(calc)
    sheet.cell(row=index, column=3).value = perclos
sheet.cell(row=len(perlist)+1, column=3).value = 0
'''

# 读取两列数据
x_values = []
y_values = []

# 假设数据从第2行开始，第1列为X轴数据，第3列为Y轴数据
for row in range(30, sheet.max_row + 1):
    x_values.append(sheet.cell(row=row, column=1).value)
    y_values.append(sheet.cell(row=row, column=3).value)

# # 绘制折线图
# plt.figure(figsize=(10, 6))
# plt.plot(x_values, y_values, marker='o', linestyle='-', color='b')
# plt.xlabel('X Axis Label')
# plt.ylabel('Y Axis Label')
# plt.title('Line Chart from Excel Data')
# plt.grid(True)

# 绘制柱状图
plt.figure(figsize=(10, 6))
plt.bar(x_values, y_values, color='b', width=0.3)
plt.xlabel('time(frame)')
plt.ylabel('perclos(%)')
# plt.title('Bar Chart from Excel Data')
plt.xticks(rotation=45)  # 旋转X轴标签，避免重叠
plt.tight_layout()  # 自动调整子图参数，使之填充整个图像区域

# 显示图表
plt.show()

# 保存修改后的Excel文件
# workbook.save(file_path)