import torch

from torch.autograd import Variable
from detection import *
from ssd_net_vgg import *
from voc0712 import *
import torch.nn as nn
import numpy as np
import cv2
import utils
import os
import openpyxl as xl

# (不支持文件，文件夹不存在会报错)
def del_files(dir_path):
    # os.walk会得到dir_path下各个后代文件夹和其中的文件的三元组列表，顺序自内而外排列，
    # 如 log下有111文件夹，111下有222文件夹：[('D:\\log\\111\\222', [], ['22.py']), ('D:\\log\\111', ['222'], ['11.py']), ('D:\\log', ['111'], ['00.py'])]
    for root, dirs, files in os.walk(dir_path, topdown=False):
        # print(root) # 各级文件夹绝对路径
        # print(dirs) # root下一级文件夹名称列表，如 ['文件夹1','文件夹2']
        # print(files)  # root下文件名列表，如 ['文件1','文件2']
        # 第一步：删除文件
        for name in files:
            os.remove(os.path.join(root, name))  # 删除文件
        # 第二步：删除空文件夹
        for name in dirs:
            os.rmdir(os.path.join(root, name)) # 删除一个空目录
        print(dir_path+'clear')

if torch.cuda.is_available():
    torch.set_default_tensor_type('torch.cuda.FloatTensor')
colors_tableau = [(255, 255, 255), (31, 119, 180), (174, 199, 232), (255, 127, 14), (255, 187, 120),
                 (44, 160, 44), (152, 223, 138), (214, 39, 40), (255, 152, 150),
                 (148, 103, 189), (197, 176, 213), (140, 86, 75), (196, 156, 148),
                 (227, 119, 194), (247, 182, 210), (127, 127, 127), (199, 199, 199),
                 (188, 189, 34), (219, 219, 141), (23, 190, 207), (158, 218, 229),(158, 218, 229),(158, 218, 229)]

net = SSD()    # initialize SSD
net = torch.nn.DataParallel(net)
net.train(mode=False)
net.load_state_dict(torch.load('./weights/ssd_voc_5000_plus.pth',map_location=lambda storage, loc: storage))
# img_id = 60
# name='6'
# image = cv2.imread('D:\\WorkSpace\\Fatigue-Driven-Detection-Based-on-CNN-master\\tests\splited\\'+name+'.jpg', cv2.IMREAD_COLOR)

# cap = cv2.VideoCapture(i_video)
file_name='D:\\WorkSpace\\Fatigue-Driven-Detection-Based-on-CNN-master\\data\\1.mp4'
cap=cv2.VideoCapture(file_name)
# VideoCapture()中的参数若为0，则表示打开笔记本的内置摄像头
# 若为视频文件路径，则表示打开视频

# del_files(r'D:\WorkSpace\Fatigue-Driven-Detection-Based-on-CNN-master\data\splited')
# del_files(r'D:\WorkSpace\Fatigue-Driven-Detection-Based-on-CNN-master\data\done')

num_frame = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) # 获取视频总帧数
print(num_frame)

expand_name = '.jpg'
if not cap.isOpened():
    print("Please check the path.")

cnt = 0 
# while 1:
#     ret, frame = cap.read()
#     # cap.read()表示按帧读取视频。ret和frame是获取cap.read()方法的两个返回值
#     # 其中，ret是布尔值。如果读取正确，则返回TRUE；如果文件读取到视频最后一帧的下一帧，则返回False
#     # frame就是每一帧的图像

#     if not ret:
#         break

#     cnt += 1 # 从1开始计帧数
#     cv2.imwrite(r'D:\WorkSpace\Fatigue-Driven-Detection-Based-on-CNN-master\data\splited\\' + str(cnt) + expand_name, frame)

# 创建一个工作簿
workbook = xl.Workbook()
sheet = workbook.active
sheet.title = "Sheet1"  # 可以设置工作表的标题
# sheet.cell(row=0, column=0).value = 'index'
# sheet.cell(row=0, column=1).value = 'eye_stat'

for index in range(1,num_frame+1):
    image = cv2.imread('D:\\WorkSpace\\Fatigue-Driven-Detection-Based-on-CNN-master\\data\\splited\\'+str(index)+'.jpg', cv2.IMREAD_COLOR)
    sheet.cell(row=index, column=1).value = index
    num_rec = 0 #检测到的眼睛的数量
    flag_B=True #是否闭眼的flag

    x = cv2.resize(image, (300, 300)).astype(np.float32)
    x -= (104.0, 117.0, 123.0)
    x = x.astype(np.float32)
    x = x[:, :, ::-1].copy()
    # plt.imshow(x)
    x = torch.from_numpy(x).permute(2, 0, 1)
    xx = Variable(x.unsqueeze(0))     # wrap tensor in Variable
    if torch.cuda.is_available():
        xx = xx.cuda()
    y = net(xx)
    softmax = nn.Softmax(dim=-1)
     # detect = Detect(config.class_num, 0, 200, 0.01, 0.45)
    detect = Detect.apply # pytorch新版本需要这样使用
    priors = utils.default_prior_box()

    loc,conf = y
    loc = torch.cat([o.view(o.size(0), -1) for o in loc], 1)
    conf = torch.cat([o.view(o.size(0), -1) for o in conf], 1)

    detections = detect(
        loc.view(loc.size(0), -1, 4),
        softmax(conf.view(conf.size(0), -1,config.class_num)),
        torch.cat([o.view(-1, 4) for o in priors], 0),
        config.class_num,
        200,
        0.7,
        0.45
    ).data
    # detections = detect.apply

    labels = VOC_CLASSES
    top_k=10

    # plt.imshow(rgb_image)  # plot the image for matplotlib

    # scale each detection back up to the image
    scale = torch.Tensor(image.shape[1::-1]).repeat(2)
    for i in range(detections.size(1)):
        j = 0
        while detections[0,i,j,0] >= 0.4:
            score = detections[0,i,j,0]
            label_name = labels[i-1]
            if label_name=='closed_eye':
                flag_B=False
            display_txt = '%s: %.2f'%(label_name, score)
            pt = (detections[0,i,j,1:]*scale).cpu().numpy()
            pt = pt.astype(int)
            coords = (pt[0], pt[1]), pt[2]-pt[0]+1, pt[3]-pt[1]+1
            color = colors_tableau[i]
            cv2.rectangle(image,(pt[0],pt[1]), (pt[2],pt[3]), color, 2)
            cv2.putText(image, display_txt, (int(pt[0]), int(pt[1]) + 10), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255,255,255), 1, 8)
            j+=1
            num_rec+=1
    if num_rec>1:
        if flag_B:
            #print(' 1:eye-open')
            # list_B=np.append(list_B,1)#睁眼为‘1’
            sheet.cell(row=index, column=2).value = 1
        else:
            #print(' 0:eye-closed')
            # list_B=np.append(list_B,0)#闭眼为‘0’
            sheet.cell(row=index, column=2).value = 0
        # list_B=np.delete(list_B,0)
    else:
        # print('nothing detected')
        sheet.cell(row=index, column=2).value = 0
    #print(list)
    # cv2.imshow('test', image)
    # cv2.waitKey(10000)
    # print("------end-------")
    cv2.imwrite('D:\\WorkSpace\\Fatigue-Driven-Detection-Based-on-CNN-master\\data\\done\\'+str(index)+'_done.jpg',image)
workbook.save('D:\\WorkSpace\\Fatigue-Driven-Detection-Based-on-CNN-master\\data\\perclos.xlsx')



print("------end-------")
'''
x = cv2.resize(image, (300, 300)).astype(np.float32)
x -= (104.0, 117.0, 123.0)
x = x.astype(np.float32)
x = x[:, :, ::-1].copy()
# plt.imshow(x)
x = torch.from_numpy(x).permute(2, 0, 1)
xx = Variable(x.unsqueeze(0))     # wrap tensor in Variable
if torch.cuda.is_available():
    xx = xx.cuda()
y = net(xx)
softmax = nn.Softmax(dim=-1)
 # detect = Detect(config.class_num, 0, 200, 0.01, 0.45)
detect = Detect.apply # pytorch新版本需要这样使用
priors = utils.default_prior_box()

loc,conf = y
loc = torch.cat([o.view(o.size(0), -1) for o in loc], 1)
conf = torch.cat([o.view(o.size(0), -1) for o in conf], 1)

detections = detect(
    loc.view(loc.size(0), -1, 4),
    softmax(conf.view(conf.size(0), -1,config.class_num)),
    torch.cat([o.view(-1, 4) for o in priors], 0),
    config.class_num,
    200,
    0.7,
    0.45
).data
# detections = detect.apply

labels = VOC_CLASSES
top_k=10

# plt.imshow(rgb_image)  # plot the image for matplotlib

# scale each detection back up to the image
scale = torch.Tensor(image.shape[1::-1]).repeat(2)
for i in range(detections.size(1)):
    j = 0
    while detections[0,i,j,0] >= 0.4:
        score = detections[0,i,j,0]
        label_name = labels[i-1]
        display_txt = '%s: %.2f'%(label_name, score)
        pt = (detections[0,i,j,1:]*scale).cpu().numpy()
        pt = pt.astype(int)
        coords = (pt[0], pt[1]), pt[2]-pt[0]+1, pt[3]-pt[1]+1
        color = colors_tableau[i]
        cv2.rectangle(image,(pt[0],pt[1]), (pt[2],pt[3]), color, 2)
        cv2.putText(image, display_txt, (int(pt[0]), int(pt[1]) + 10), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255,255,255), 1, 8)
        j+=1
cv2.imshow('test', image)
cv2.waitKey(10000)
print("------end-------")
cv2.imwrite('./tests/'+name+'_done.jpg',image)
'''